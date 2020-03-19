from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

import csv
import pprint

    # 余分な文字列の削除
def replace_str(string):
    string = string.replace('\r\n', '')
    string = string.replace(' ', '')
    string = string.replace(u'\xa0', '')
    string = string.replace(u'\u3000', '')
    string = string.replace('\n', '')
    return string

class Webscraping:


    def __init__(self,filename):
        self.filename = filename
        self.soup = BeautifulSoup(open(self.filename),"html.parser")
        self.soup_without_table = BeautifulSoup(open(self.filename),"html.parser")
        self.Chapter_Nums = list()
        self.Chapter_Contents = list()
        self.Tables = list()

    
    def ExtractChapterNum(self,soup):
        Chapter_Nums = soup.find_all(["h1","h2","h3"])

        for Chapter_num in Chapter_Nums:
            self.Chapter_Nums.append(replace_str(Chapter_num.get_text()))

    def ExtractChapterContent(self,soup):
        Chapter_Contents = soup.find_all(["h1","h2","h3","p"])

        for Chapter_Content in Chapter_Contents:
            self.Chapter_Contents.append(replace_str(Chapter_Content.get_text()))
    
    def ExtractTables(self,soup):
        Tables = soup.findAll("table", {"class":"MsoNormalTable"})

        for Table in Tables:
            self.Tables.append(Table)
    
    def TableSize(self):
        return len(self.Tables)

    def Remove_tags(self,soup):
        for tag in soup.findAll("table"):
            # タグとその内容の削除
            tag.decompose()



webscraping = Webscraping('test.html')
webscraping.ExtractChapterNum(webscraping.soup)
webscraping.ExtractTables(webscraping.soup)


webscraping.Remove_tags(webscraping.soup_without_table)
webscraping.ExtractChapterContent(webscraping.soup_without_table)

webscraping.Chapter_Contents = [a for a in webscraping.Chapter_Contents if a != '']

#print(webscraping.Chapter_Nums)

#print(webscraping.Chapter_Contents)

#print(webscraping.Tables)

# テーブルを指定



#################################################

wb = openpyxl.Workbook()
sheet  = wb.active

count = 0
for i, value in enumerate(webscraping.Chapter_Contents):

    if (value in webscraping.Chapter_Nums):

        #章番号を書き込み
        sheet.cell(row=i+5-count, column=4).value = value
        count +=1
    else:
        #本文を書き込み
        sheet.cell(row=i+5-count, column=5).value = value


sheat_tables = list()
for i in range(0,webscraping.TableSize()):
    sheat_tables.append(wb.create_sheet())


# 保存
wb.save('example.xlsx')


tableMatrix = []
for table in webscraping.Tables:
    #Here you can do whatever you want with the data! You can findAll table row headers, etc...
    list_of_rows = []
    for row in table.findAll('tr')[0:]:
        list_of_cells = []
        for cell in row.findAll('td'):
            text = cell.text.replace('&nbsp;', '')
            list_of_cells.append(text)
        list_of_rows.append(list_of_cells)
    tableMatrix.append((list_of_rows, list_of_cells))
pprint.pprint(tableMatrix[3], width=40)